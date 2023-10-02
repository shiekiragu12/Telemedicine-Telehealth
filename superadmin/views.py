from django.contrib import messages
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import render, redirect

from facilities.utils import is_not_blank_or_empty
from mailer.models import Email, EmailConfiguration
from facilities.models import Doctor, Patient, Staff, County, Constituency, Service, ServiceCategory, Condition, \
    Speciality, Facility, FacilityType, Appointment, DoctorNote, Illness
from mailer.utils import send_custom_email
from shop.models import Category, Product, Order, ProductType
from mainapp.models import Blog, Tag, AppConfig, Contact
from website.models import Book, Schedule, Analytic, Apply, Equip, Telehealth, DemoRequest, Firstaid
from .decorators import admin_only
import openpyxl
from .models import Media
from mailer.moreviews import send_doctor_authorized_email


# Create your views here.


@admin_only
def dashboard(request):
    users_ = User.objects.all()
    doctors_ = Doctor.objects.all()
    staff_ = Staff.objects.all()
    patients_ = Patient.objects.all()
    facilities_ = Facility.objects.all()
    facility_types_ = FacilityType.objects.all()
    products_ = Product.objects.all()
    specialities_ = Speciality.objects.all()
    diseases_ = Condition.objects.all()
    blogs_ = Blog.objects.all()

    context = {
        "users_count": users_.count(),
        "users": users_.order_by('-id')[0:5],
        "doctors_count": doctors_.count(),
        "staff_count": staff_.count(),
        "patients_count": patients_.count(),
        "facilities": facilities_.order_by('-id')[0:5],
        "clinics_count": facilities_.filter(facility_kind="clinic").count(),
        "pharmacies_count": facilities_.filter(facility_kind='pharmacy').count(),
        "nutraceutacals_count": facilities_.filter(facility_kind='nutraceuticals').count(),
        "facility_types_count": facility_types_.count(),
        "products": products_.order_by('id')[0:5],
        "live_products_count": products_.filter(status='approved').count(),
        "pending_products_count": products_.filter(status='pending').count(),
        "specialities_count": specialities_.count(),
        "diseases_count": diseases_.count(),
        "blogs": blogs_.order_by('-id')[0:5],
        "total_blogs": blogs_.count()
    }
    return render(request, template_name='super-admin-dashboard/pages/index.html', context=context)


# App settings
@admin_only
def settings(request):
    context = {
        "app": AppConfig.objects.filter(app='main').first(),
        "emails": Email.objects.all(),
        "emailconfigs": EmailConfiguration.objects.all(),
    }
    return render(request, template_name='super-admin-dashboard/pages/settings/settings.html', context=context)


# Admin Notifications
@admin_only
def admin_notifications(request):
    context = {
        "app": AppConfig.objects.filter(app='main').first(),
        "emails": Email.objects.all(),
        "emailconfigs": EmailConfiguration.objects.all(),
    }
    return render(request, template_name='super-admin-dashboard/pages/settings/notifications.html', context=context)


# Users
@admin_only
def users(request):
    users_ = User.objects.get_queryset().order_by('id')
    users_2 = User.objects.get_queryset().order_by('id')

    gender = request.GET.get('gender', None)

    search = request.GET.get('search', None)
    ordering = request.GET.get('ordering', None)
    limit = request.GET.get('limit')  # Show 25 results per page.

    if is_not_blank_or_empty(gender):
        users_2 = users_2.filter(profile__gender=gender)

    if is_not_blank_or_empty(search):
        users_2 = users_2.filter(
            Q(username__icontains=search) |
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(email__icontains=search))

    if is_not_blank_or_empty(ordering):
        users_2 = users_2.order_by(ordering)

    paginator = Paginator(users_2, limit if is_not_blank_or_empty(limit) else 10)  # Show 25 facilities per page.

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number if int(page_number) <= paginator.num_pages else 1)

    context = {
        "page_obj": page_obj,
        "pages": paginator.page_range,
        "page_count": paginator.num_pages,
        "paginator_count": paginator.count,
        "total_users": users_.count(),
        "total_admins": users_.filter(is_superuser=True).count(),
        "active_users": users_.filter(is_active=True).count(),
    }

    return render(request, template_name='super-admin-dashboard/pages/users/users.html', context=context)


@admin_only
def single_user(request, user_id):
    context = {
        'user': User.objects.filter(id=user_id).first()
    }
    return render(request, template_name='super-admin-dashboard/pages/users/single-user.html', context=context)


@admin_only
def doctors(request):
    doctors_ = Doctor.objects.get_queryset().order_by('id')
    doctors_2 = Doctor.objects.get_queryset().order_by('id')

    gender = request.GET.get('gender', None)
    speciality = request.GET.get('speciality', None)
    authorized = request.GET.get('authorized', None)

    search = request.GET.get('search', None)
    ordering = request.GET.get('ordering', None)
    limit = request.GET.get('limit')  # Show 25 results per page.

    if is_not_blank_or_empty(speciality):
        doctors_2 = doctors_2.filter(specialities__id=speciality)

    if is_not_blank_or_empty(gender):
        doctors_2 = doctors_2.filter(user__profile__gender=gender)

    if is_not_blank_or_empty(authorized):
        doctors_2 = doctors_2.filter(is_verified=authorized)

    if is_not_blank_or_empty(search):
        doctors_2 = doctors_2.filter(
            Q(user__username__icontains=search) |
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__email__icontains=search) |
            Q(license_number__icontains=search) |
            Q(regulatory_body__icontains=search) |
            Q(about__icontains=search))

    if is_not_blank_or_empty(ordering):
        doctors_2 = doctors_2.order_by(ordering)

    paginator = Paginator(doctors_2, limit if is_not_blank_or_empty(limit) else 25)  # Show 25 facilities per page.

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number if int(page_number) <= paginator.num_pages else 1)

    context = {
        "page_obj": page_obj,
        "pages": paginator.page_range,
        "page_count": paginator.num_pages,
        "paginator_count": paginator.count,
        "total_doctors": doctors_.count(),
    }
    return render(request, template_name='super-admin-dashboard/pages/users/doctors/doctors.html', context=context)


@admin_only
def single_doctor(request, doctor_id):
    context = {
        'doctor': Doctor.objects.filter(id=doctor_id).first()
    }
    return render(request, template_name='super-admin-dashboard/pages/users/doctors/single-doctor.html',
                  context=context)


@admin_only
def change_doctor_status(request, doctor_id, status):
    doctor = Doctor.objects.filter(id=doctor_id).first()
    if doctor:
        if status == 'True':
            doctor.is_verified = True
            doctor.save()
            messages.info(request, 'Doctor has been Authorized/Verified')
            # send_doctor_authorized_email(doctor)
            send_custom_email('doctor_authorized', doctor, [doctor.user.email])
        else:
            doctor.is_verified = False
            doctor.save()
            messages.info(request, 'Doctor has been De-Authorized')
    else:
        messages.error(request, 'Doctor not found')
    return redirect(request.META['HTTP_REFERER'])


@admin_only
def staff(request):
    staff_ = Staff.objects.get_queryset().order_by('id')
    staff_2 = Staff.objects.get_queryset().order_by('id')

    gender = request.GET.get('gender', None)

    search = request.GET.get('search', None)
    ordering = request.GET.get('ordering', None)
    limit = request.GET.get('limit')  # Show 25 results per page.

    if is_not_blank_or_empty(gender):
        staff_2 = staff_2.filter(user__profile__gender=gender)

    if is_not_blank_or_empty(search):
        staff_2 = staff_2.filter(
            Q(user__username__icontains=search) |
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__email__icontains=search) |
            Q(designation__icontains=search) |
            Q(education__icontains=search))

    if is_not_blank_or_empty(ordering):
        staff_2 = staff_2.order_by(ordering)

    paginator = Paginator(staff_2, limit if is_not_blank_or_empty(limit) else 10)  # Show 25 facilities per page.

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number if int(page_number) <= paginator.num_pages else 1)

    context = {
        "page_obj": page_obj,
        "pages": paginator.page_range,
        "page_count": paginator.num_pages,
        "paginator_count": paginator.count,
        "total_staff": staff_.count(),
    }
    return render(request, template_name='super-admin-dashboard/pages/users/staff/staff.html', context=context)


@admin_only
def single_staff(request, staff_id):
    context = {
        'staff': Staff.objects.filter(id=staff_id).first()
    }
    return render(request, template_name='super-admin-dashboard/pages/users/staff/single-staff.html', context=context)


@admin_only
def patients(request):
    patients_ = Patient.objects.get_queryset().order_by('id')
    patients_2 = Patient.objects.get_queryset().order_by('id')

    gender = request.GET.get('gender', None)
    blood_group = request.GET.get('blood_group', None)

    search = request.GET.get('search', None)
    ordering = request.GET.get('ordering', None)
    limit = request.GET.get('limit')  # Show 25 results per page.

    if is_not_blank_or_empty(gender):
        patients_2 = patients_2.filter(user__profile__gender=gender)

    if is_not_blank_or_empty(blood_group):
        patients_2 = patients_2.filter(blood_group__icontains=blood_group)

    if is_not_blank_or_empty(search):
        patients_2 = patients_2.filter(
            Q(user__username__icontains=search) |
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__email__icontains=search))

    if is_not_blank_or_empty(ordering):
        patients_2 = patients_2.order_by(ordering)

    paginator = Paginator(patients_2, limit if is_not_blank_or_empty(limit) else 10)  # Show 25 facilities per page.

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number if int(page_number) <= paginator.num_pages else 1)

    context = {
        "page_obj": page_obj,
        "pages": paginator.page_range,
        "page_count": paginator.num_pages,
        "paginator_count": paginator.count,
        "total_patients": patients_.count(),
    }
    return render(request, template_name='super-admin-dashboard/pages/users/patients/patients.html', context=context)


@admin_only
def single_patient(request, patient_id):
    context = {
        'patient': Patient.objects.filter(id=patient_id).first()
    }
    return render(request, template_name='super-admin-dashboard/pages/users/patients/single-patient.html',
                  context=context)


# Emails
@admin_only
def emails(request):
    emails_ = Email.objects.all()
    context = {
        "emails": emails_,
        "total_emails": emails_.count(),
    }
    return render(request, template_name='super-admin-dashboard/pages/emails/emails.html', context=context)


@admin_only
def emailconfigs(request):
    emailconfigs_ = EmailConfiguration.objects.all()
    context = {
        "emailconfigs": emailconfigs_,
        "total_configs": emailconfigs_.count(),
    }
    return render(request, template_name='super-admin-dashboard/pages/emails/emailconfigs.html', context=context)


@admin_only
def email_variables(request):
    return render(request, template_name='super-admin-dashboard/pages/emails/variables.html', context={})


# Locations
@admin_only
def counties(request):
    if request.method == "POST":
        counties_input = request.POST.get('counties')
        counties_list = counties_input.split(',')
        counties_to_save = [County(name=mystr.strip()) for mystr in counties_list]
        County.objects.bulk_create(counties_to_save)
        messages.success(request, 'Counties created successfully')

    counties_ = County.objects.get_queryset().order_by('id')
    counties_2 = County.objects.get_queryset().order_by('id')

    search = request.GET.get('search', None)
    ordering = request.GET.get('ordering', None)
    limit = request.GET.get('limit')  # Show 25 results per page.

    paginator = Paginator(counties_2,
                          limit if is_not_blank_or_empty(limit) else 25)  # Show 25 facilities per page.

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number if int(page_number) <= paginator.num_pages else 1)

    context = {
        "page_obj": page_obj,
        "pages": paginator.page_range,
        "page_count": paginator.num_pages,
        "paginator_count": paginator.count,
        "total_counties": counties_.count(),
    }

    return render(request, template_name='super-admin-dashboard/pages/locations/counties.html', context=context)


@admin_only
def constituencies(request):
    if request.method == "POST":
        county = County.objects.filter(id=request.POST.get('county'))
        constituency_ = request.POST.get('const')

        const_ = Constituency(county=county, name=constituency_)
        const_.save()

        messages.success(request, 'Constituency created successfully')

    constituencies_ = Constituency.objects.get_queryset().order_by('id')
    constituencies_2 = Constituency.objects.get_queryset().order_by('id')

    search = request.GET.get('search', None)
    ordering = request.GET.get('ordering', None)
    limit = request.GET.get('limit')  # Show 25 results per page.

    paginator = Paginator(constituencies_2,
                          limit if is_not_blank_or_empty(limit) else 25)  # Show 25 facilities per page.

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number if int(page_number) <= paginator.num_pages else 1)

    context = {
        "page_obj": page_obj,
        "pages": paginator.page_range,
        "page_count": paginator.num_pages,
        "paginator_count": paginator.count,
        "total_constituencies": constituencies_.count(),
    }

    return render(request, template_name='super-admin-dashboard/pages/locations/constituencies.html', context=context)


# Facilities
@admin_only
def conditions(request):
    if request.method == "POST":
        conditions_input = request.POST.get('conditions')
        conditions_list = conditions_input.split(',')
        conditions_to_save = [Condition(name=mystr.strip()) for mystr in conditions_list]
        Condition.objects.bulk_create(conditions_to_save)
        messages.success(request, 'Conditions created successfully')
        return redirect(request.META.get('HTTP_REFERER'))

    conditions_ = Condition.objects.get_queryset().order_by('id')
    conditions_2 = Condition.objects.get_queryset().order_by('id')

    search = request.GET.get('search', None)
    ordering = request.GET.get('ordering', None)
    limit = request.GET.get('limit')  # Show 25 results per page.

    paginator = Paginator(conditions_2, limit if is_not_blank_or_empty(limit) else 25)  # Show 25 facilities per page.

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number if int(page_number) <= paginator.num_pages else 1)

    context = {
        "page_obj": page_obj,
        "pages": paginator.page_range,
        "page_count": paginator.num_pages,
        "paginator_count": paginator.count,
        "total_conditions": conditions_.count(),
    }

    return render(request, template_name='super-admin-dashboard/pages/facilities/conditions.html', context=context)


@admin_only
def illness_page(request):
    illness_ = Illness.objects.all()
    if request.method == "POST":
        conditions_input = request.POST.get('illness')
        conditions_list = conditions_input.split(',')
        conditions_to_save = [Illness(name=mystr.strip()) for mystr in conditions_list]
        Illness.objects.bulk_create(conditions_to_save)
        messages.success(request, 'Illness created successfully')
        return redirect(request.META.get('HTTP_REFERER'))

    illness_ = Illness.objects.get_queryset().order_by('id')
    illness_2 = Illness.objects.get_queryset().order_by('id')

    search = request.GET.get('search', None)
    ordering = request.GET.get('ordering', None)
    limit = request.GET.get('limit')  # Show 25 results per page.

    paginator = Paginator(illness_2, limit if is_not_blank_or_empty(limit) else 25)  # Show 25 facilities per page.

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number if int(page_number) <= paginator.num_pages else 1)

    context = {
        "page_obj": page_obj,
        "pages": paginator.page_range,
        "page_count": paginator.num_pages,
        "paginator_count": paginator.count,
        "total_illness": illness_.count(),
    }

    return render(request, template_name='super-admin-dashboard/pages/facilities/illness.html', context=context)


@admin_only
def specialities(request):
    specialities_ = Speciality.objects.all()
    if request.method == "POST":
        specialities_input = request.POST.get('specialities')
        specialities_list = specialities_input.split(',')
        specialities_to_save = [Speciality(name=mystr.strip()) for mystr in specialities_list]
        Speciality.objects.bulk_create(specialities_to_save)
        messages.success(request, 'Specialities created successfully')
        return redirect(request.META.get('HTTP_REFERER'))
    context = {
        "specialities": specialities_,
        "total_specialities": specialities_.count(),
    }
    return render(request, template_name='super-admin-dashboard/pages/facilities/specialities.html', context=context)


@admin_only
def service_categories(request):
    if request.method == "POST":
        services_input = request.POST.get('services')
        services_list = services_input.split(',')
        services_to_save = [ServiceCategory(name=mystr.strip()) for mystr in services_list]
        ServiceCategory.objects.bulk_create(services_to_save)
        messages.success(request, 'Service Categories created successfully')
        return redirect(request.META.get('HTTP_REFERER'))

    service_categories_ = ServiceCategory.objects.get_queryset().order_by('id')
    service_categories_2 = ServiceCategory.objects.get_queryset().order_by('id')

    search = request.GET.get('search', None)
    ordering = request.GET.get('ordering', None)
    limit = request.GET.get('limit')  # Show 25 results per page.

    paginator = Paginator(service_categories_2,
                          limit if is_not_blank_or_empty(limit) else 25)  # Show 25 facilities per page.

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number if int(page_number) <= paginator.num_pages else 1)

    context = {
        "page_obj": page_obj,
        "pages": paginator.page_range,
        "page_count": paginator.num_pages,
        "paginator_count": paginator.count,
        "total_service_categories": service_categories_.count(),
    }
    return render(request, template_name='super-admin-dashboard/pages/facilities/services_categories.html',
                  context=context)


@admin_only
def services(request):
    services_ = Service.objects.get_queryset().order_by('id')
    services_2 = Service.objects.get_queryset().order_by('id')

    search = request.GET.get('search', None)
    ordering = request.GET.get('ordering', None)
    limit = request.GET.get('limit')  # Show 25 results per page.

    paginator = Paginator(services_2,
                          limit if is_not_blank_or_empty(limit) else 25)  # Show 25 facilities per page.

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number if int(page_number) <= paginator.num_pages else 1)

    context = {
        "page_obj": page_obj,
        "pages": paginator.page_range,
        "page_count": paginator.num_pages,
        "paginator_count": paginator.count,
        "total_services": services_.count(),
    }
    return render(request, template_name='super-admin-dashboard/pages/facilities/services.html', context=context)


@admin_only
def facilities(request):
    context = facility_model_reader(request, 'all')
    return render(request, template_name='super-admin-dashboard/pages/facilities/facilities.html', context=context)


@admin_only
def single_facility(request, facility_id):
    context = {
        "facility": Facility.objects.filter(id=facility_id).first(),
    }
    return render(request, template_name='super-admin-dashboard/pages/facilities/single-facility.html', context=context)


@admin_only
def hospitals(request):
    context = facility_model_reader(request, 'hospital')
    return render(request, template_name='super-admin-dashboard/pages/facilities/hospitals/hospitals.html',
                  context=context)


@admin_only
def clinics(request):
    context = facility_model_reader(request, 'clinic')
    return render(request, template_name='super-admin-dashboard/pages/facilities/clinics/clinics.html', context=context)


@admin_only
def pharmacies(request):
    context = facility_model_reader(request, 'pharmacy')
    return render(request, template_name='super-admin-dashboard/pages/facilities/pharmacies/pharmacies.html',
                  context=context)


@admin_only
def nutraceuticals(request):
    context = facility_model_reader(request, 'nutraceutical')
    return render(request, template_name='super-admin-dashboard/pages/facilities/nutras/nutras.html', context=context)


@admin_only
def change_facility_status(request, facility_id, status):
    facility = Facility.objects.filter(id=facility_id).first()
    if facility:
        if status == 'True':
            facility.authorized = True
            facility.save()
            # send_facility_authorized_email(facility)
            messages.info(request, 'Facility Authorized successfully')
            send_custom_email('facility_authorized', facility, [facility.email,
                                                                facility.owner.email if facility.owner else ''])
        else:
            facility.authorized = False
            facility.save()
            messages.info(request, 'Facility De-Authorized successfully')

    else:
        messages.error(request, 'Facility not found')
    return redirect(request.META['HTTP_REFERER'])


@admin_only
def facility_types(request):
    if request.method == "POST":
        types_input = request.POST.get('types')
        types_list = types_input.split(',')
        types_to_save = [FacilityType(name=mystr.strip()) for mystr in types_list]
        FacilityType.objects.bulk_create(types_to_save)
        messages.success(request, 'Facility Types created successfully')
        return redirect(request.META.get('HTTP_REFERER'))

    facility_types_ = FacilityType.objects.get_queryset().order_by('id')
    facility_types_2 = ServiceCategory.objects.get_queryset().order_by('id')

    search = request.GET.get('search', None)
    ordering = request.GET.get('ordering', None)
    limit = request.GET.get('limit')  # Show 25 results per page.

    paginator = Paginator(facility_types_2,
                          limit if is_not_blank_or_empty(limit) else 25)  # Show 25 facilities per page.

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number if int(page_number) <= paginator.num_pages else 1)

    context = {
        "page_obj": page_obj,
        "pages": paginator.page_range,
        "page_count": paginator.num_pages,
        "paginator_count": paginator.count,
        "total_facility_types": facility_types_.count(),
    }

    return render(request, template_name='super-admin-dashboard/pages/facilities/facility_types.html', context=context)


@admin_only
def appointments(request):
    appointments_ = Appointment.objects.get_queryset().order_by('id')
    appointments_2 = Appointment.objects.get_queryset().order_by('id')

    facility = request.GET.get('facility', None)
    doctor = request.GET.get('doctor', None)
    patient = request.GET.get('patient', None)

    status = request.GET.get('status', None)

    search = request.GET.get('search', None)
    ordering = request.GET.get('ordering', None)
    limit = request.GET.get('limit')  # Show 25 results per page.

    if is_not_blank_or_empty(facility):
        appointments_2 = appointments_2.filter(Q(facility__facility_code__icontains=facility) |
                                               Q(facility__name__icontains=facility))

    if is_not_blank_or_empty(doctor):
        appointments_2 = appointments_2.filter(Q(doctor__user__email__icontains=facility) |
                                               Q(doctor__user__first_name__icontains=facility))

    if is_not_blank_or_empty(patient):
        appointments_2 = appointments_2.filter(Q(patient__user__email__icontains=facility) |
                                               Q(patient__user__first_name__icontains=facility))

    if is_not_blank_or_empty(status):
        appointments_2 = appointments_2.filter(status=status)

    if is_not_blank_or_empty(search):
        appointments_2 = appointments_2.filter(
            Q(note__icontains=search))

    if is_not_blank_or_empty(ordering):
        staff_2 = appointments_2.order_by(ordering)

    paginator = Paginator(appointments_2, limit if is_not_blank_or_empty(limit) else 10)  # Show 25 facilities per page.

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number if int(page_number) <= paginator.num_pages else 1)

    context = {
        "page_obj": page_obj,
        "pages": paginator.page_range,
        "page_count": paginator.num_pages,
        "paginator_count": paginator.count,
        "total_staff": appointments_.count(),
    }
    return render(request, template_name='super-admin-dashboard/pages/facilities/appointments.html', context=context)


@admin_only
def encounters(request):
    encounters_ = DoctorNote.objects.get_queryset().order_by('id')
    encounters_2 = DoctorNote.objects.get_queryset().order_by('id')

    search = request.GET.get('search', None)
    ordering = request.GET.get('ordering', None)
    limit = request.GET.get('limit')  # Show 25 results per page.

    paginator = Paginator(encounters_2, limit if is_not_blank_or_empty(limit) else 25)  # Show 25 facilities per page.

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number if int(page_number) <= paginator.num_pages else 1)

    context = {
        "page_obj": page_obj,
        "pages": paginator.page_range,
        "page_count": paginator.num_pages,
        "paginator_count": paginator.count,
        "total_encounters": encounters_.count(),
    }
    return render(request, template_name='super-admin-dashboard/pages/facilities/encounters.html', context=context)


# Shop
@admin_only
def product_categories(request):
    if request.method == "POST":
        categories_input = request.POST.get('categories')
        categories_list = categories_input.split(',')
        categories_to_save = [Category(name=tag.strip()) for tag in categories_list]
        Category.objects.bulk_create(categories_to_save)
        messages.success(request, 'Categories created successfully')
        return redirect(request.META.get('HTTP_REFERER'))

    product_categories_ = Category.objects.get_queryset().order_by('id')
    product_categories_2 = Category.objects.get_queryset().order_by('id')

    search = request.GET.get('search', None)
    ordering = request.GET.get('ordering', None)
    limit = request.GET.get('limit')  # Show 25 results per page.

    paginator = Paginator(product_categories_2,
                          limit if is_not_blank_or_empty(limit) else 25)  # Show 25 facilities per page.

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number if int(page_number) <= paginator.num_pages else 1)

    context = {
        "page_obj": page_obj,
        "pages": paginator.page_range,
        "page_count": paginator.num_pages,
        "paginator_count": paginator.count,
        "total_product_categories": product_categories_.count(),
    }

    return render(request, template_name='super-admin-dashboard/pages/shop/product_categories.html', context=context)


@admin_only
def product_types(request):
    if request.method == "POST":
        product_types_input = request.POST.get('types')
        product_type_list = product_types_input.split(',')
        product_types_to_save = [ProductType(name=tag.strip()) for tag in product_type_list]
        ProductType.objects.bulk_create(product_types_to_save)
        messages.success(request, 'Product types created successfully')
        return redirect(request.META.get('HTTP_REFERER'))

    product_types_ = ProductType.objects.get_queryset().order_by('id')
    product_types_2 = ProductType.objects.get_queryset().order_by('id')

    search = request.GET.get('search', None)
    ordering = request.GET.get('ordering', None)
    limit = request.GET.get('limit')  # Show 25 results per page.

    paginator = Paginator(product_types_2,
                          limit if is_not_blank_or_empty(limit) else 25)  # Show 25 facilities per page.

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number if int(page_number) <= paginator.num_pages else 1)

    context = {
        "page_obj": page_obj,
        "pages": paginator.page_range,
        "page_count": paginator.num_pages,
        "paginator_count": paginator.count,
        "total": product_types_.count(),
    }

    return render(request, template_name='super-admin-dashboard/pages/shop/product_types.html', context=context)


@admin_only
def products(request):
    products_ = Product.objects.get_queryset().order_by('id')
    products_2 = Product.objects.get_queryset().order_by('id')

    facility = request.GET.get('facility', None)
    available = request.GET.get('available', None)
    product_type = request.GET.get('product_type', None)
    category = request.GET.get('category', None)
    tags_ = request.GET.get('tags', None)
    status = request.GET.get('status', None)

    search = request.GET.get('search', None)
    ordering = request.GET.get('ordering', None)

    if is_not_blank_or_empty(facility):
        products_2 = products_2.filter(facility__facility_code=facility)

    if is_not_blank_or_empty(available):
        products_2 = products_2.filter(available=available)

    if is_not_blank_or_empty(product_type):
        products_2 = products_2.filter(product_type=product_type)

    if is_not_blank_or_empty(category):
        products_2 = products_2.filter(categories__id__in=[int(category)])

    if is_not_blank_or_empty(tags_):
        products_2 = products_2.filter(tags__id__in=[int(tags_)])

    if is_not_blank_or_empty(status):
        products_2 = products_2.filter(status=status)

    if is_not_blank_or_empty(search):
        products_2 = products_2.filter(
            Q(name__icontains=search) | Q(code__icontains=search) | Q(description__icontains=search) | Q(
                dom__icontains=search))

    if is_not_blank_or_empty(ordering):
        products_2 = products_2.order_by(ordering)

    paginator = Paginator(products_2, 25)  # Show 25 products per page.

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number if int(page_number) <= paginator.num_pages else 1)

    context = {
        "page_obj": page_obj,
        "pages": paginator.page_range,
        "page_count": paginator.num_pages,
        "paginator_count": paginator.count,
        "total_products": products_.count(),
    }
    return render(request, template_name='super-admin-dashboard/pages/shop/products.html', context=context)


@admin_only
def single_product(request, product_id):
    context = {
        'product': Product.objects.filter(id=product_id).first()
    }
    return render(request, template_name='super-admin-dashboard/pages/shop/single-product.html', context=context)


@admin_only
def change_product_status(request, product_id, status):
    product = Product.objects.filter(id=product_id).first()
    if product:
        product.status = status
        product.save()
        messages.success(request, "Product updated successfully")
    else:
        messages.error(request, 'Product not found')
    return redirect(request.META['HTTP_REFERER'])


@admin_only
def orders(request):
    orders_ = Order.objects.all()
    context = {
        "orders": orders_,
        "total_orders": orders_.count(),
    }
    orders_ = Order.objects.get_queryset().order_by('id')
    orders_2 = Order.objects.get_queryset().order_by('id')

    search = request.GET.get('search', None)
    ordering = request.GET.get('ordering', None)
    limit = request.GET.get('limit')  # Show 25 results per page.

    if is_not_blank_or_empty(search):
        orders_2 = orders_2.filter(
            Q(user__first_name__icontains=search) | Q(user__last_name__icontains=search) |
            Q(user__username__icontains=search) |
            Q(user__email__icontains=search))

    paginator = Paginator(orders_2,
                          limit if is_not_blank_or_empty(limit) else 25)  # Show 25 facilities per page.

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number if int(page_number) <= paginator.num_pages else 1)

    context = {
        "page_obj": page_obj,
        "pages": paginator.page_range,
        "page_count": paginator.num_pages,
        "paginator_count": paginator.count,
        "total_orders": orders_.count(),
    }
    return render(request, template_name='super-admin-dashboard/pages/shop/orders.html', context=context)


@admin_only
def single_order(request, order_id):
    context = {
        'order': Order.objects.filter(id=order_id).first()
    }
    return render(request, template_name='super-admin-dashboard/pages/shop/single-order.html', context=context)


def blogs(request):
    blogs_ = Blog.objects.all()
    context = {
        "blogs": blogs_,
        "total_blogs": blogs_.count(),
    }
    return render(request, template_name='super-admin-dashboard/pages/blogs/blogadmin.html', context=context)


def new_blog(request):
    context = {"tags": Tag.objects.all()}
    if request.method == 'POST':
        tags_ = Tag.objects.filter(id__in=request.POST.getlist('tags'))

        blog = Blog.objects.create(
            created_by=request.user,
            title=request.POST.get('title'),
            body=request.POST.get('body'),
            image=request.FILES.get('image'),
            is_public=True if request.POST.get('is_public', True) == 'on' else False,
        )
        blog.tags.set(tags_)
        blog.save()
        messages.success(request, "Blog published successfully")
        return redirect('super-admin-new-blog')
    return render(request, template_name='super-admin-dashboard/pages/blogs/new-blog.html', context=context)


def update_blog(request, blog_id):
    blog = Blog.objects.get(id=blog_id)
    context = {
        "tags": Tag.objects.all(),
        "blog": blog
    }
    if request.method == 'POST':
        tags_ = Tag.objects.filter(id__in=request.POST.getlist('tags'))
        blog.tags.set(tags_)
        blog.title = request.POST.get('title')
        blog.body = request.POST.get('body')
        blog.is_public = True if request.POST.get('is_public', True) == 'on' else False
        image = request.FILES.get('image', None)
        if image:
            blog.image = image
        blog.save()
        messages.success(request, "Blog updated successfully")
        return redirect(request.META['HTTP_REFERER'])
    return render(request, template_name='super-admin-dashboard/pages/blogs/update-blog.html', context=context)


@admin_only
def upload_conditions(request):
    if request.method == 'POST':
        excel_file = request.FILES.get("file")
        wb = openpyxl.load_workbook(excel_file)
        excel_data = list()

        for letter in ["A"]:
            worksheet = wb[letter]
            for row in worksheet.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data.append(row_data)

        conditions_ = []
        for condition_data in excel_data:
            available = Condition.objects.filter(name=condition_data[0].strip()).first()
            if available is None:
                cond = Condition(name=condition_data[0].strip(),
                                 factsheet=condition_data[1].strip(),
                                 pathogen=condition_data[2].strip(),
                                 clinical_features=condition_data[3].strip(),
                                 transmission=condition_data[4].strip(),
                                 diagnosis=condition_data[5].strip(),
                                 treatment=condition_data[6].strip(),
                                 prevention=condition_data[7].strip())
                conditions_.append(cond)
        Condition.objects.bulk_create(conditions_)

    return render(request, template_name='super-admin-dashboard/pages/conditions/upload-conditions.html', context={})


@admin_only
def upload_specialities(request):
    if request.method == 'POST':
        excel_file = request.FILES.get("file")
        wb = openpyxl.load_workbook(excel_file)
        excel_data = list()

        for letter in ["A"]:
            worksheet = wb[letter]
            for row in worksheet.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data.append(row_data)

        specialities_ = []
        for data in excel_data:
            available = Speciality.objects.filter(name=data[0].strip()).first()
            if available is None:
                cond = Speciality(name=data[0].strip())
                specialities_.append(cond)
        Speciality.objects.bulk_create(specialities_)
        messages.success(request, "Specialities uploaded successfully")

    return render(request, template_name='super-admin-dashboard/pages/specialities/upload-specialities.html',
                  context={})


@admin_only
def upload_illness(request):
    if request.method == 'POST':
        excel_file = request.FILES.get("file")
        wb = openpyxl.load_workbook(excel_file)
        excel_data = list()

        for letter in ["A"]:
            worksheet = wb[letter]
            for row in worksheet.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data.append(row_data)

        illness_ = []
        for data in excel_data:
            available = Illness.objects.filter(name=data[0]).first()
            if available is None:
                cond = Illness(name=data[0])
                illness_.append(cond)
        Illness.objects.bulk_create(illness_)
        messages.success(request, "Illness uploaded successfully")

    return render(request, template_name='super-admin-dashboard/pages/illness/upload-illness.html', context={})


def get_county(county_name):
    county = County.objects.filter(name=county_name.strip().lower()).first()
    if county:
        return county
    else:
        county_ = County.objects.create(name=county_name.strip().lower())
        return county_


def get_constituency(const_name, county_name):
    const_ = Constituency.objects.filter(name=const_name.strip().lower()).first()
    if const_:
        return const_
    else:
        county_ = get_county(county_name)
        const__ = Constituency.objects.create(name=const_name.strip().lower(), county=county_)
        return const__


def get_facility_type(type_name):
    f_type = FacilityType.objects.filter(name=type_name.strip().lower()).first()
    if f_type:
        return f_type
    else:
        f_type_ = FacilityType.objects.create(name=type_name.strip().lower())
        return f_type_


@admin_only
def upload_medical_facilities(request):
    # Facility.objects.all().delete()
    # County.objects.all().delete()
    # Constituency.objects.all().delete()
    if request.method == 'POST':
        excel_file = request.FILES.get("file")
        wb = openpyxl.load_workbook(excel_file)
        excel_data = list()

        for county in ["A", "B", "C", "D", "F", "G", "H"]:
            worksheet = wb[county]
            for i, row in enumerate(worksheet.iter_rows()):
                if i != 0:
                    row_data = list()
                    for cell in row:
                        row_data.append(str(cell.value))
                    excel_data.append(row_data)

        facilities_ = []
        for facility_data in excel_data:

            available = Facility.objects.filter(name=facility_data[0].strip()).first()

            if available is None:
                facility_ = Facility(owner_name=facility_data[2].strip(),
                                     facility_type=get_facility_type(facility_data[1]),
                                     name=facility_data[0],
                                     county=get_county(facility_data[3]),
                                     constituency=get_constituency(facility_data[4], facility_data[3]),
                                     latitude=facility_data[5],
                                     longitude=facility_data[6],
                                     email=facility_data[7],
                                     contact_no=facility_data[8])
                facilities_.append(facility_)
        # Bulk created the Facilities
        Facility.objects.bulk_create(facilities_)
        return redirect('super-admin-upload-facilities')

    return render(request, template_name='super-admin-dashboard/pages/facilities/upload-facilities.html', context={})


def tags(request):
    tags_ = Tag.objects.all()
    if request.method == "POST":
        tags_input = request.POST.get('tags')
        tags_list = tags_input.split(',')
        tags_to_save = [Tag(name=tag.strip()) for tag in tags_list]
        Tag.objects.bulk_create(tags_to_save)
        messages.success(request, 'Tags created successfully')
        return redirect(request.META.get('HTTP_REFERER'))

    context = {
        "tags": tags_,
        "total_tags": tags_.count(),
    }
    return render(request, template_name='super-admin-dashboard/pages/blogs/tags.html', context=context)


def media(request):
    media_ = Media.objects.all()
    if request.method == "POST":
        media_input = request.FILES.getlist('media')
        for file in media_input:
            Media.objects.create(file=file)
        messages.success(request, 'Media uploaded successfully')
        return redirect(request.META.get('HTTP_REFERER'))

    context = {
        "photos": media_,
        "total_media": media_.count(),
    }
    return render(request, template_name='super-admin-dashboard/pages/media/media.html', context=context)


# Contact forms/All contact messages via different forms
@admin_only
def contact_form(request):
    context = form_model_reader(request, Contact)

    return render(request, template_name='super-admin-dashboard/pages/contact/contact-form.html', context=context)


@admin_only
def book_form(request):
    context = form_model_reader(request, Book)
    return render(request, template_name='super-admin-dashboard/pages/contact/book-form.html', context=context)


@admin_only
def schedule_form(request):
    context = form_model_reader(request, Schedule)
    return render(request, template_name='super-admin-dashboard/pages/contact/schedule-form.html', context=context)


@admin_only
def analytic_form(request):
    context = form_model_reader(request, Analytic)
    return render(request, template_name='super-admin-dashboard/pages/contact/analytic-form.html', context=context)


@admin_only
def apply_form(request):
    context = form_model_reader(request, Apply)
    return render(request, template_name='super-admin-dashboard/pages/contact/apply-form.html', context=context)


@admin_only
def equip_form(request):
    context = form_model_reader(request, Equip)
    return render(request, template_name='super-admin-dashboard/pages/contact/equip-form.html', context=context)


@admin_only
def telehealth_form(request):
    context = form_model_reader(request, Telehealth)
    return render(request, template_name='super-admin-dashboard/pages/contact/telehealth-form.html', context=context)


@admin_only
def top_button_form(request):
    context = form_model_reader(request, DemoRequest)
    return render(request, template_name='super-admin-dashboard/pages/contact/topbutton-form.html', context=context)


@admin_only
def first_aid_form(request):
    context = form_model_reader(request, Firstaid)
    return render(request, template_name='super-admin-dashboard/pages/contact/firstaid-form.html', context=context)


def form_model_reader(request, model):
    objs_ = model.objects.get_queryset().order_by('id')
    objs_2 = model.objects.get_queryset().order_by('id')

    read = request.GET.get('read', None)
    search = request.GET.get('search', None)
    ordering = request.GET.get('ordering', None)
    limit = request.GET.get('limit')  # Show 25 results per page.

    if is_not_blank_or_empty(read):
        objs_2 = objs_2.filter(read=read)

    paginator = Paginator(objs_2,
                          limit if is_not_blank_or_empty(limit) else 25)  # Show 25 facilities per page.

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number if int(page_number) <= paginator.num_pages else 1)

    context = {
        "page_obj": page_obj,
        "pages": paginator.page_range,
        "page_count": paginator.num_pages,
        "paginator_count": paginator.count,
        "total": objs_.count(),
        "read": objs_.filter(read=True).count(),
        "unread": objs_.filter(read=False).count(),
    }
    return context


def facility_model_reader(request, facility_kind):
    facilities_ = Facility.objects.get_queryset().filter(facility_kind=facility_kind).order_by('id')
    facilities_2 = Facility.objects.get_queryset().filter(facility_kind=facility_kind).order_by('id')

    if facility_kind == "all":
        facilities_ = Facility.objects.get_queryset().order_by('id')
        facilities_2 = Facility.objects.get_queryset().order_by('id')

    county = request.GET.get('county', None)
    constituency = request.GET.get('constituency', None)
    authorized = request.GET.get('authorized', None)
    search = request.GET.get('search', None)
    ordering = request.GET.get('ordering', None)

    if is_not_blank_or_empty(county):
        facilities_2 = facilities_2.filter(county__id=county)

    if is_not_blank_or_empty(constituency):
        facilities_2 = facilities_2.filter(constituency__id=constituency)

    if is_not_blank_or_empty(authorized):
        facilities_2 = facilities_2.filter(authorized=authorized)

    if is_not_blank_or_empty(search):
        facilities_2 = facilities_2.filter(
            Q(owner_name__icontains=search) | Q(facility_code__icontains=search) | Q(name__icontains=search) | Q(
                description__icontains=search) | Q(email__icontains=search) | Q(contact_no__icontains=search) | Q(
                address__icontains=search) | Q(home_page_content__icontains=search) | Q(
                about_page_content__icontains=search) | Q(online_page_content__icontains=search))

    if is_not_blank_or_empty(ordering):
        facilities_2 = facilities_2.order_by(ordering)

    paginator = Paginator(facilities_2, 25)  # Show 25 facilities per page.

    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number if int(page_number) <= paginator.num_pages else 1)

    context = {
        "page_obj": page_obj,
        "pages": paginator.page_range,
        "page_count": paginator.num_pages,
        "paginator_count": paginator.count,
        "total_facilities": facilities_.count(),
    }

    return context
